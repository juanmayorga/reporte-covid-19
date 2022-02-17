import requests
import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from openpyxl.chart import (
    LineChart,
    Reference,
    BarChart,
)
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.axis import DateAxis
import lxml.html as html


HOME_URL = 'https://github.com/MinCiencia/Datos-COVID19/blob/master/output/producto1/Covid-19.csv'
HOME_URL_SHEET2 = 'https://github.com/MinCiencia/Datos-COVID19/blob/master/output/producto90/incidencia_en_vacunados.csv'

XPATH_HEADERS = '//tr[@id="LC1"]/th/text()'
XPATH_HEADERS_SHEET2 = '//tr[@id="LC1"]/th/text()'


def home():
    try:
        response = requests.get(HOME_URL)

        if response.status_code == 200:
            home = response.content.decode('utf-8')
            parsed = html.fromstring(home)
            comuna = input('Ingresa comuna: ')
            if comuna == 'Arica':
                comunaxpath = '//tr[@id="LC2"]/td/text()'
            elif comuna == 'Antofagasta':
                comunaxpath = '//tr[@id="LC15"]/td/text()'
            elif comuna == 'Coquimbo':
                comunaxpath = '//tr[@id="LC38"]/td/text()'
            elif comuna == 'Valparaiso':
                comunaxpath = '//tr[@id="LC85"]/td/text()'
            elif comuna == 'Maule':
                comunaxpath = '//tr[@id="LC188"]/td/text()'
            elif comuna == "Talca":
                comunaxpath = '//tr[@id="LC202"]/td/text()'
            elif comuna == 'Aysén' or comuna == 'Aysen':
                comunaxpath = '//tr[@id="LC341"]/td/text()'
            else:
                comunaxpath = '//tr[contains(.,"' + comuna + '")]/td/text()'

            headers = parsed.xpath(XPATH_HEADERS)
            data = parsed.xpath(comunaxpath)
            today = datetime.date.today().strftime('%d-%m-%Y')
            wb = Workbook()
            ws = wb.active
            ws.title = "reporte " + today
            ws['A1'] = 'Estos datos son obtenidos desde el repositorio Github del Ministerio de Ciencias: https://github.com/MinCiencia/Datos-COVID19/blob/master/output/producto1/Covid-19.csv'
            for i in range(0, 5):
                _ = ws.cell(column=i+1, row=3, value=headers[i])
                _ = ws.cell(column=i+1, row=4, value=data[i])

            ws['A6'] = 'Fecha'
            ws['B6'] = 'Casos Totales'
            ws['C6'] = 'Casos Diarios'

            # fechas
            # for i in range(5, len(data)-1):
            #   _ = ws.cell(column=1, row=i, value=headers[i])
            # casos Totales
            for i in range(7, len(data)-1):
                _ = ws.cell(column=1, row=i, value=headers[i])
                _ = ws.cell(column=2, row=i, value=int(float(data[i])))
            # casos diarios
            for i in range(7, len(data)-1):
                if i == 7:
                    _ = ws.cell(column=3, row=i, value='=B' +
                                str(i))
                else:
                    _ = ws.cell(column=3, row=i, value='=B' +
                                str(i)+'-B'+str(i-1))

            for i in range(1, 6):
                ws.column_dimensions[get_column_letter(i)].auto_size = True

            chart = BarChart()
            chart.title = "Casos diarios para la comuna de " + comuna
            chart.type = "col"
            chart.style = 13
            chart.y_axis.title = "Casos diarios"
            chart.x_axis.title = "Fecha"
            chart.x_axis.number_format = 'dd-mm-yyyy'
            chart.height = 15
            chart.width = 45

            values = Reference(ws, min_col=3, min_row=4,
                               max_col=3, max_row=len(data)-2)

            chart.add_data(values, titles_from_data=True)

            dates = Reference(ws, min_col=1, min_row=5,
                              max_col=1, max_row=len(data)-2)
            chart.set_categories(dates)
            chart.legend.layout = Layout(
                manualLayout=ManualLayout(
                    yMode='edge',
                    xMode='edge',
                    x=1, y=0.1,
                    h=0.1, w=1
                )
            )
            ws.add_chart(chart, "E6")

            #wb.save(f'Reporte {comuna} {today}.xlsx')

        else:
            raise ValueError(f'Error: {response.status_code}')

        response = requests.get(HOME_URL_SHEET2)

        if response.status_code == 200:
            home = response.content.decode('utf-8')
            parsed = html.fromstring(home)
            ws2 = wb.create_sheet('Incidencia en vacunados')
            ws2['A1'] = 'Estos datos son obtenidos desde el repositorio Github del Ministerio de Ciencias: https://github.com/MinCiencia/Datos-COVID19/blob/master/output/producto90/incidencia_en_vacunados.csv'

            #response = requests.get(HOME_URL_SHEET2)

            headers_sheet2 = parsed.xpath(XPATH_HEADERS_SHEET2)

            ws2['A3'] = headers_sheet2[0]
            ws2['D3'] = 'Fallecidos sin vacunas'
            ws2['G3'] = 'Fallecidos con vacunas (1,2,3 o 4 dosis)'
        else:
            raise ValueError(f'Error: {response.status_code}')

        wb.save(f'Reporte {comuna} {today}.xlsx')

    except ValueError as e:
        print(e)


def run():
    home()


if __name__ == '__main__':
    run()
